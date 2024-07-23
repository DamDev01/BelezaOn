from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from .models import Cliente, Servico, Agendamento
from django.http import HttpResponseRedirect
from django.urls import reverse
from .forms import ServicoForm, AgendamentoForm
from django.urls import reverse_lazy
from django.contrib import messages
from .forms import AgendamentoForm
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from django.db.models import Sum
from django.http import JsonResponse
from .models import Servico
from django.shortcuts import redirect
from itertools import groupby
from collections import defaultdict
import calendar
from .models import Produto, Categoria
from .forms import ProdutoForm, AtualizarQuantidadeForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
import datetime
from django.views import View
import json
from decimal import Decimal
from django.contrib.auth import authenticate, login
from .forms import LoginForm
from django.contrib.auth import logout


@login_required
def index(request):
    agendamentos = Agendamento.objects.all()
    return render(request, 'salon/index.html', {'agendamentos': agendamentos})

@login_required
def novo_agendamento(request):
    if request.method == 'POST':
        nome_cliente = request.POST['nome_cliente']
        servico_id = request.POST['servico']
        data = request.POST['data']
        hora = request.POST['hora']
        valor = request.POST['valor']
        
        # Obtendo o cliente
        cliente, created = Cliente.objects.get_or_create(nome=nome_cliente)
        
        # Verificando se o serviço existe
        servico = get_object_or_404(Servico, pk=servico_id)
        
        # Criando o agendamento
        agendamento = Agendamento.objects.create(cliente=cliente, servico=servico, data=data, hora=hora, valor=valor)
        
        # Redirecionando para a página de sucesso
        messages.success(request, 'Agendamento criado com sucesso!')
        return HttpResponseRedirect(reverse('home'))
    else:
        servicos = Servico.objects.all()
        return render(request, 'salon/novo_agendamento.html', {'servicos': servicos})

@login_required
def editar_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    
    if request.method == 'POST':
        form = AgendamentoForm(request.POST, instance=agendamento)
        if form.is_valid():
            # Salvando o agendamento sem modificar o nome do cliente
            form.save(commit=False)
            form.cliente = agendamento.cliente  # Mantém o cliente original
            form.save()
            messages.success(request, 'Agendamento atualizado com sucesso!')
            return redirect('agendamentos_pendentes')
    else:
        # Passando o nome do cliente como valor inicial para garantir que ele seja preenchido no formulário
        form = AgendamentoForm(instance=agendamento, initial={'nome_cliente': agendamento.cliente.nome})
    
    servicos = Servico.objects.all()
    return render(request, 'salon/editar_agendamento.html', {'form': form, 'servicos': servicos})


@login_required
def excluir_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == 'POST':
        agendamento.delete()
        return redirect('agendamentos_pendentes')
    return render(request, 'salon/excluir_agendamento.html', {'agendamento': agendamento})

@login_required
def agendamentos_pendentes(request):
    agendamentos = Agendamento.objects.filter(concluido=False)
    for agendamento in agendamentos:
        agendamento.hora_formatada = agendamento.hora.strftime("%H:%M")
    return render(request, 'salon/agendamentos_pendentes.html', {'agendamentos': agendamentos})


@login_required
def concluir_agendamento(request, agendamento_id):
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == 'POST':
        forma_pagamento = request.POST.get('forma_pagamento')
        agendamento.concluido = True
        agendamento.forma_pagamento = forma_pagamento
        agendamento.save()
        return redirect('agendamentos_pendentes')
    return render(request, 'salon/agendamentos_pendentes.html')

@login_required
def agendamentos_concluidos(request):
    today = timezone.now()
    current_year = today.year

    # Obtendo o parâmetro de filtro de forma de pagamento
    forma_pagamento = request.GET.get('forma_pagamento', '')

    # Obtendo todos os agendamentos concluídos do ano atual com o filtro de forma de pagamento
    if forma_pagamento:
        agendamentos = Agendamento.objects.filter(concluido=True, data__year=current_year, forma_pagamento=forma_pagamento)
    else:
        agendamentos = Agendamento.objects.filter(concluido=True, data__year=current_year)

    # Criando um dicionário de tradução para os meses em português
    meses_pt_br = {i: calendar.month_name[i] for i in range(1, 13)}

    # Agrupando agendamentos por mês
    agendamentos_por_mes = defaultdict(list)
    for agendamento in agendamentos:
        if agendamento.data:  # Verifica se a data não é None
            month = meses_pt_br[agendamento.data.month]
            agendamento.hora_formatada = agendamento.hora.strftime("%H:%M") if agendamento.hora else ""
            agendamentos_por_mes[month].append(agendamento)

    # Criando uma estrutura de dados para o template
    agendamentos_data = []
    for month, agendamentos_list in agendamentos_por_mes.items():
        valor_total = sum(float(agendamento.valor) for agendamento in agendamentos_list)
        agendamentos_data.append({
            'month': month,
            'agendamentos': agendamentos_list,
            'valor_total': valor_total
        })

    return render(request, 'salon/agendamentos_concluidos.html', {
        'agendamentos_data': agendamentos_data,
        'forma_pagamento_selecionada': forma_pagamento,
    })

@login_required
def lista_servicos(request):
    servicos = Servico.objects.all()
    return render(request, 'salon/lista_servicos.html', {'servicos': servicos})

@login_required
def editar_servico(request, servico_id):
    servico = get_object_or_404(Servico, pk=servico_id)
    if request.method == 'POST':
        form = ServicoForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            return redirect('lista_servicos')
    else:
        form = ServicoForm(instance=servico)
    return render(request, 'salon/editar_servico.html', {'form': form})

@login_required
def excluir_servico(request, servico_id):
    servico = get_object_or_404(Servico, pk=servico_id)
    if request.method == 'POST':
        servico.delete()
        return redirect('lista_servicos')
    return render(request, 'salon/excluir_servico.html', {'servico': servico})

@login_required
def agendar_servico(request):
    if request.method == 'POST':
        form = AgendamentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('página_de_sucesso')
    else:
        form = AgendamentoForm()
    return render(request, 'agendamento_form.html', {'form': form})

@login_required
def home(request):
    return render(request, 'salon/home.html')

@login_required   
def obter_valor_servico(request):
    if request.method == 'GET' and 'servico_id' in request.GET:
        servico_id = request.GET['servico_id']
        try:
            servico = Servico.objects.get(pk=servico_id)
            return JsonResponse({'valor': servico.valor})
        except Servico.DoesNotExist:
            return JsonResponse({'error': 'Serviço não encontrado'}, status=404)
    return JsonResponse({'error': 'Método inválido'}, status=400)

@login_required
def dashboard(request):
    return render(request, 'dashboard.html')

@login_required
def adicionar_produto(request):
    if request.method == 'POST':
        form = ProdutoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_produtos')
    else:
        form = ProdutoForm()
    return render(request, 'salon/adicionar_produto.html', {'form': form})

@login_required
def lista_produtos(request):
    produtos = Produto.objects.all()
    categorias = Categoria.objects.all()
    return render(request, 'salon/lista_produtos.html', {'produtos': produtos, 'categorias': categorias})

@login_required
def atualizar_quantidade(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    if request.method == 'POST':
        form = AtualizarQuantidadeForm(request.POST, instance=produto)
        if form.is_valid():
            form.save()
            return redirect('lista_produtos')
    else:
        form = AtualizarQuantidadeForm(instance=produto)
    return render(request, 'salon/atualizar_quantidade.html', {'form': form, 'produto': produto})

@login_required
def filtrar_por_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    produtos = Produto.objects.filter(categoria=categoria)
    categorias = Categoria.objects.all()
    return render(request, 'salon/lista_produtos.html', {'produtos': produtos, 'categorias': categorias, 'categoria_selecionada': categoria})

@login_required
def pesquisar_cliente(request):
    query = request.GET.get('cliente')
    resultados = []

    if query:
        resultados = Agendamento.objects.filter(cliente__nome__icontains=query, concluido=True)
        total_gasto = resultados.aggregate(Sum('valor'))['valor__sum'] or 0
    else:
        total_gasto = 0

    context = {
        'resultados': resultados,
        'total_gasto': total_gasto,
        'query': query
    }

    return render(request, 'salon/pesquisar_cliente.html', context)

class AgendamentosRelatorioView(View):
    def get(self, request):
        forma_pagamento = request.GET.get('forma_pagamento', '')
        agendamentos = Agendamento.objects.filter(concluido=True)

        if forma_pagamento:
            agendamentos = agendamentos.filter(forma_pagamento=forma_pagamento)

        agendamentos_data = []
        current_month = None
        mes_data = {'month': '', 'agendamentos': [], 'valor_total': 0}

        for agendamento in agendamentos.order_by('data'):
            agendamento_mes = agendamento.data.strftime('%B %Y')
            if agendamento_mes != current_month:
                if current_month:
                    agendamentos_data.append(mes_data)
                current_month = agendamento_mes
                mes_data = {'month': current_month, 'agendamentos': [], 'valor_total': 0}

            mes_data['agendamentos'].append(agendamento)
            mes_data['valor_total'] += agendamento.valor

        if current_month:
            agendamentos_data.append(mes_data)

        return render(request, 'salon/relatorio_agendamentos.html', {
            'agendamentos_data': agendamentos_data,
            'forma_pagamento_selecionada': forma_pagamento
        })

@login_required
def exportar_agendamentos_excel(request):
    forma_pagamento = request.GET.get('forma_pagamento', '')
    agendamentos = Agendamento.objects.filter(concluido=True)

    if forma_pagamento:
        agendamentos = agendamentos.filter(forma_pagamento=forma_pagamento)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Agendamentos Concluídos'

    # Estilos
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Cabeçalho
    columns = ['Mês', 'Nome do Cliente', 'Serviço', 'Data', 'Hora', 'Valor', 'Forma de Pagamento']
    worksheet.append(columns)
    for cell in worksheet[1]:
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.fill = yellow_fill
        cell.border = border

    # Dados
    current_month = None
    total_valor_mes = 0

    for agendamento in agendamentos.order_by('data'):
        agendamento_mes = f'{meses[agendamento.data.month]} {agendamento.data.year}'
        if agendamento_mes != current_month:
            if current_month:
                worksheet.append(['', '', '', '', 'Total', f'R${total_valor_mes}'])
                for cell in worksheet[worksheet.max_row]:
                    cell.font = font_bold
                    cell.alignment = alignment_center
                    cell.border = border
                for cell in worksheet[f'A{worksheet.max_row}:G{worksheet.max_row}'][0]:
                    cell.fill = yellow_fill
            current_month = agendamento_mes
            total_valor_mes = 0

        total_valor_mes += agendamento.valor
        worksheet.append([
            agendamento_mes,
            agendamento.cliente.nome,
            agendamento.servico.nome,
            agendamento.data.strftime('%d/%m/%Y'),
            agendamento.hora.strftime('%H:%M'),
            f'R${agendamento.valor}',
            agendamento.forma_pagamento.capitalize()
        ])
        for cell in worksheet[worksheet.max_row]:
            cell.border = border

    if current_month:
        worksheet.append(['', '', '', '', 'Total', f'R${total_valor_mes}'])
        for cell in worksheet[worksheet.max_row]:
            cell.font = font_bold
            cell.alignment = alignment_center
            cell.border = border
        for cell in worksheet[f'A{worksheet.max_row}:G{worksheet.max_row}'][0]:
            cell.fill = yellow_fill

    # Ajustar largura das colunas
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Relatorio_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    workbook.save(response)
    return response

# Dicionário para mapear os meses em português
meses = {
    1: 'Janeiro',
    2: 'Fevereiro',
    3: 'Março',
    4: 'Abril',
    5: 'Maio',
    6: 'Junho',
    7: 'Julho',
    8: 'Agosto',
    9: 'Setembro',
    10: 'Outubro',
    11: 'Novembro',
    12: 'Dezembro'
}

@login_required
def decimal_to_float(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError

@login_required
def agendamentos_graficos(request):
    forma_pagamento_selecionada = request.GET.get('forma_pagamento', '')
    agendamentos = Agendamento.objects.all()
    
    if forma_pagamento_selecionada:
        agendamentos = agendamentos.filter(forma_pagamento=forma_pagamento_selecionada)
    
    agendamentos_data = defaultdict(lambda: {'agendamentos': [], 'valor_total': 0})
    formas_pagamento = defaultdict(int)
    clientes_por_mes = defaultdict(int)
    
    for agendamento in agendamentos:
        mes_ano = agendamento.data.strftime('%Y-%m')
        agendamentos_data[mes_ano]['agendamentos'].append(agendamento)
        agendamentos_data[mes_ano]['valor_total'] += float(agendamento.valor)  # Convertendo para float
        formas_pagamento[agendamento.forma_pagamento] += 1
        clientes_por_mes[mes_ano] += 1
    
    agendamentos_data = [
        {
            'month': calendar.month_name[int(m.split('-')[1])],
            'year': m.split('-')[0],
            'valor_total': data['valor_total']
        }
        for m, data in sorted(agendamentos_data.items())
    ]
    
    clientes_por_mes_data = [
        {
            'month': calendar.month_name[int(m.split('-')[1])],
            'year': m.split('-')[0],
            'count': count
        }
        for m, count in sorted(clientes_por_mes.items())
    ]

    context = {
        'agendamentos_data_json': json.dumps(agendamentos_data, default=decimal_to_float),
        'forma_pagamento_selecionada': forma_pagamento_selecionada,
        'formas_pagamento_json': json.dumps(dict(formas_pagamento), default=decimal_to_float),
        'clientes_por_mes_json': json.dumps(clientes_por_mes_data, default=decimal_to_float)
    }
    return render(request, 'salon/agendamentos_graficos.html', context)


@login_required
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirecionar para uma página de sucesso após o login
                return redirect('home.html')
            else:
                # Se a autenticação falhar, exibir mensagem de erro no formulário
                form.add_error(None, 'Usuário ou senha inválidos. Tente novamente.')
    else:
        form = LoginForm()
    return render(request, 'home.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "Você saiu do sistema.")
    return render(request, 'logout.html')